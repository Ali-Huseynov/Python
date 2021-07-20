from django.shortcuts import render, HttpResponse
import pandas as pd
from openpyxl import load_workbook


df = pd.read_excel( "Section_4.xlsx" , engine="openpyxl" )

def convert_ipv4(ip):
    return tuple(int(n) for n in ip.split('.'))

def check_ipv4_in(addr, start, end):
    return convert_ipv4(start) <= convert_ipv4(addr) <= convert_ipv4(end)

def search( x, addr ):
    ip_range = ( x["ip_from"] , x["tip_to"] )
    return check_ipv4_in( addr , *ip_range)

def ip_is_valid(ip):

    if ip.replace(".","").isnumeric():
        return all( [ (int(i) <=255) for i in ip.split(".") ])
    return False


def home(request):
    context = {}
    if request.method == "POST":
        ip = request.POST["ip_name"]
        if ip_is_valid(ip):
            d = df[ df.apply( search , axis=1, addr =ip  ) ]
            if d.empty :
                context["not_found"] = "No matching range"
            else:
                context["addresses"] = zip( d["ip_from"] , d["tip_to"]  )
        else:
            context["error"] = "IP address is not valid"
        
    return render(request, "ip_checker/home.html", context=context )

#------------------- ADD RANGE  ---------------------------

def range_is_valid(ip_from,ip_to):

    return ( ( ip_is_valid(ip_from) and ip_is_valid(ip_to) ) and (convert_ipv4(ip_from) < convert_ipv4(ip_to)) )

def search_range(x,addr):
    ip_from, ip_to = addr
    
    return ( check_ipv4_in( ip_from,x["ip_from"], x["tip_to"]  ) or 
            check_ipv4_in(ip_to,x["ip_from"], x["tip_to"]) or 
            check_ipv4_in( x["ip_from"],ip_from,ip_to  ) or 
            check_ipv4_in(x["tip_to"], ip_from,ip_to) 
    )


def save_range(ip_from, ip_to):
    global df
    add_df =  pd.DataFrame( { "ip_from":ip_from, "tip_to":ip_to } ,index=[ df.shape[0]+1 ] )

    file_name = "Section_4.xlsx"
    writer = pd.ExcelWriter(file_name, engine='openpyxl', mode='a',)

    writer.book = load_workbook(file_name)
        
    startrow = writer.book["Sheet1"].max_row

    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    add_df.to_excel(writer, "Sheet1", index=False, startrow=startrow, header=False)

    writer.save()
    df = pd.concat( [ df  , add_df ] )


def add_range(request):
    context = {}
    if request.method == "POST":
        ip_from = request.POST["ip_from"]
        ip_to = request.POST["ip_to"]

        if range_is_valid(ip_from,ip_to) :
            d = df[ df.apply( search_range , axis=1, addr = ( ip_from,ip_to)  ) ]
            
            if (d.empty)  :
                save_range( ip_from,ip_to )
                context["success"] = "successfully added new range"
            else:
                context["exist"] = { "msg": "That range already exist", "data":zip( d["ip_from"] , d["tip_to"]  ) }
        else:
            context["error"] = "IP address is not valid"
    
    return render(request, "ip_checker/add_range.html", context=context )